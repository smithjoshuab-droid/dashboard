"""
Reads the loan spreadsheet and rebuilds index.html from the template.
Runs inside GitHub Actions after download_spreadsheet.py.
"""
import os, sys, json, datetime
import openpyxl

# ── Helpers ──────────────────────────────────────────────────────────────────

def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str) and "/" in v:
        try:
            p = v.split("/")
            return f"{p[2]}-{p[0].zfill(2)}-{p[1].zfill(2)}"
        except Exception:
            pass
    return None

def to_num(v):
    try: return float(v)
    except Exception: return None

def fp_str(v):
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%d")
    return str(v).strip() if v else "N/A"

# ── Sheet readers ─────────────────────────────────────────────────────────────

def read_funded(wb, sheet):
    if sheet not in wb.sheetnames:
        print(f"  Warning: sheet '{sheet}' not found, skipping.")
        return []
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        d = dict(zip(headers, row))
        amt = to_num(d.get("Total Loan Amount"))
        if not amt: continue
        rows.append({
            "borrower":     str(d.get("Borrower","") or "").strip(),
            "loan_officer": str(d.get("Loan Officer","") or "").strip(),
            "amount":       amt,
            "fast_pass":    fp_str(d.get("Fast Pass")),
            "lender":       str(d.get("Lender","") or "").strip(),
            "purpose":      str(d.get("Purpose","") or "").strip(),
            "loan_type":    str(d.get("Loan Type","") or "").strip(),
            "funded_date":  parse_date(d.get("Funded Date")),
            "rate":         to_num(d.get("Interest Rate")),
        })
    return rows

def read_pipeline(wb):
    sheet = "Loan Pipeline"
    if sheet not in wb.sheetnames:
        print(f"  Warning: sheet '{sheet}' not found.")
        return []
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        b = row[0]
        if not b or not isinstance(b, str) or len(b.strip()) < 3: continue
        d = dict(zip(headers, row))
        amt = to_num(d.get("Total Loan Amount"))
        if not amt: continue
        rows.append({
            "borrower":      str(b).strip(),
            "loan_officer":  str(d.get("Loan Officer","") or "").strip(),
            "amount":        amt,
            "fast_pass":     fp_str(d.get("Fast Pass")),
            "lender":        str(d.get("Lender","") or "").strip(),
            "purpose":       str(d.get("Purpose","") or "").strip(),
            "loan_type":     str(d.get("Loan Type","") or "").strip(),
            "contract_close": parse_date(d.get("Contract Close Date")),
            "actual_close":  parse_date(d.get("Actual Close Date")),
            "funded_date":   parse_date(d.get("Funded Date")),
            "rate":          to_num(d.get("Interest Rate")),
        })
    return rows

# ── Main ──────────────────────────────────────────────────────────────────────

SPREADSHEET = "spreadsheet.xlsx"
TEMPLATE    = "dashboard_template.html"
OUTPUT      = "index.html"

if not os.path.exists(SPREADSHEET):
    print(f"ERROR: {SPREADSHEET} not found. Run download_spreadsheet.py first.")
    sys.exit(1)

if not os.path.exists(TEMPLATE):
    print(f"ERROR: {TEMPLATE} not found.")
    sys.exit(1)

print("Opening spreadsheet...")
wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)

data = {
    "pipeline":   read_pipeline(wb),
    "funded2026": read_funded(wb, "Apex Funded 2026"),
    "funded2025": read_funded(wb, "Apex Funded 2025"),
    "refreshed":  datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p"),
}

print(f"  pipeline:   {len(data['pipeline'])} loans")
print(f"  funded2026: {len(data['funded2026'])} loans")
print(f"  funded2025: {len(data['funded2025'])} loans")

print("Reading template...")
with open(TEMPLATE, "r", encoding="utf-8") as f:
    template = f.read()

if "%%LOAN_DATA%%" not in template:
    print("ERROR: Template is missing the %%LOAN_DATA%% placeholder.")
    sys.exit(1)

data_js = f"const RAW = {json.dumps(data, separators=(',',':'))};"
html = template.replace("%%LOAN_DATA%%", data_js)

# Inject refresh timestamp into the top bar
html = html.replace(
    'APEX<span>.</span>Mortgage',
    f'APEX<span>.</span>Mortgage <span style="font-size:11px;font-weight:400;color:var(--muted);margin-left:8px">Updated {data["refreshed"]}</span>',
    1,
)

with open(OUTPUT, "w", encoding="utf-8") as f:
    f.write(html)

print(f"Dashboard written to {OUTPUT} ({len(html):,} bytes)")
print("Done.")
